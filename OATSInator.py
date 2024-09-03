import time
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import warnings

# Remove the warning about not being able to honor data validation rules from the log
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Read in the data for the OATS
df = pd.read_excel('./Enter-OATS.xlsx', sheet_name='Data')

# Add the index to the dataframe as a column to use a reference on how to mark of the data
df.reset_index(inplace=True)

# Only select entries that haven't been entered into Orion
dfToCreate = df[df['Done'].isna()]

# Exit, if there is no new OATS entries
if(len(dfToCreate) == 0):
    exit('No new OATS entry detected in the Excel spreadsheet.')

# Split the data into Account and Opportunity level OATS
dfAccounts = dfToCreate[dfToCreate["OATS-Type"] == 'Account'] 
dfOpps = dfToCreate[dfToCreate["OATS-Type"] == 'Opportunity']

# Reset the index to make iteration easier
dfAccounts = dfAccounts.reset_index()
dfOpps = dfOpps.reset_index()

# Open up the Excel-Sheet to manipulate the data and recreate the Data Validation
fileName = './Enter-OATS.xlsx'
wb = load_workbook(fileName)
dataSheet =  wb['Data']
listSheet = wb['List']

# Orion URL
baselineURL = 'https://orion.sas.com/'

# Recreating the Data Validation in the Excel-Sheet
def saveExcel():
    # OATS-Type Validation
    dv1 = DataValidation(type="list", formula1='=Metadata!$A$2:$A$3', allow_blank=False)
    dv1.add('A2:A1048576')
    dataSheet.add_data_validation(dv1)
    listSheet.add_data_validation(dv1)

    # ID-Type Validation
    dv2 = DataValidation(type="list", formula1='=List!$D$2:$D$1048576', allow_blank=False)
    dv2.add('B2:B1048576')
    dataSheet.add_data_validation(dv2)

    # Duration-Type Validation
    dv3 = DataValidation(type="list", formula1='=Metadata!$B$2:$B$38', allow_blank=False)
    dv3.add('D2:D1048576')
    dataSheet.add_data_validation(dv3)

    # Category-Type Validation
    dv4 = DataValidation(type="list", formula1='=Metadata!$C$2:$C$23', allow_blank=False)
    dv4.add('E2:E1048576')
    dataSheet.add_data_validation(dv4)

    # Value-Proposition-Type Validation
    dv5 = DataValidation(type="list", formula1='=Metadata!$D$2:$D$64', allow_blank=False)
    dv5.add('F2:F1048576')
    dataSheet.add_data_validation(dv5)

    # Checkbox-Type Validation
    # G = Customer Facing Virtual
    # H = Customer Facing Onsite
    # I = Demonstration Given
    # J = Partner Present
    # K = Interregional Collab
    dv6 = DataValidation(type="list", formula1='=Metadata!$E$2:$E$3', allow_blank=False)
    dv6.add('G2:G1048576')
    dv6.add('H2:H1048576')
    dv6.add('I2:I1048576')
    dv6.add('J2:J1048576')
    dv6.add('K2:K1048576')
    dataSheet.add_data_validation(dv6)

    # Save the Excel-File
    wb.save(fileName)

# Open Chrome
options = webdriver.ChromeOptions()
options.add_argument("--lang=en")
# Maximize the window
options.add_argument("start-maximized")
# Remove pointless log entries from console log
options.add_experimental_option('excludeSwitches', ['enable-logging'])
# Remove installation log from console
os.environ['WDM_LOG_LEVEL'] = '0'
# Always install the current Chrome driver compatible to the installed Chrome version
print('Currently getting the new ChromeDriver...')
print('If you get an error, please enter the URL: chrome://settings/help in Chrome and update it.')
try:
    driver = webdriver.Chrome(options=options)
except:
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager()), options=options)
    except:
        try:
            driver = webdriver.Chrome(service=Service(ChromeDriverManager(driver_version='114.0.5735.90').install()), options=options)
        except:
            print('Please contact David.Weik@sas.com with your error message')

driver.get(baselineURL)
time.sleep(10)

print('Starting to process your OATS...')
print('Please note that there will be a summary at the end, containing all entries and DQ rule violations.')
print('Please also note that you are still responsible to regular check the DQ Report')

# As the Duration is bound to non searchable options, the elements position is provided
durationDictionary = {
    '15m': 1,
    '30m': 2,
    '1h': 3,
    '1h 30m': 4,
    '2h': 5,
    '2h 30m': 6,
    '3h': 7,
    '3h 30m': 8,
    '4h': 9,
    '4h 30m': 10,
    '5h': 11,
    '5h 30m': 12,
    '6h': 13,
    '6h 30m': 14,
    '7h': 15,
    '7h 30m': 16,
    '8h': 17,
    '8h 30m': 18,
    '9h': 19,
    '9h 30m': 20,
    '10h': 21,
    '10h 30m': 22,
    '11h': 23,
    '11h 30m': 24,
    '12h': 25,
    '12h 30m': 26,
    '13h': 27,
    '13h 30m': 28,
    '14h': 29,
    '14h 30m': 30,
    '15h': 31,
    '15h 30m': 32,
    '16h': 33,
    '16h 30m': 34,
    '17h': 35,
    '17h 30m': 36,
    '18h': 37,
}

# Inform user about how many Account level OATS will be entered
if(len(dfAccounts.index) == 0):
    print('No OATS for accounts to enter.')
else:
    print(f"Adding {len(dfAccounts)} account OATS.")

summary = []

def dataQualityCheck(type, entryData, entryID):
    print(f"Applying {type} DQ checks to {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']}")
    listOfDummyAccounts = ['12356240', '12356250', '12356260', '12356270', '12356280', '13049830', '13049840', '13049850']
    categoryType = entryData['Category'][entryData['Category'].find('-') + 2:]
    categoryTypeDummy = entryData['Category'].find('DUMMY')
    if(categoryTypeDummy > -1):
        categoryType = 'ACCT'
    dqCheckValue = True
    nonCustomerFacingCategories = [' DUMMY - Receiving Training - ACCT', 'Competitive Research - OPP', 'Demonstration Development - OPP/ACCT', 'Internal meeting / Review - OPP', 'Partner Meeting, Planning / Review - OPP/ACCT', 'Travel to Customer Site - OPP/ACCT']
    mustCustomerFacingCategories = ['Architecture Workshop / Meeting - OPP', 'Functional Workshop / Meeting - OPP']
    if((entryData['Customer Facing Virtual'] == 'Yes' or entryData['Customer Facing Onsite'] == 'Yes') and entryData['Category'] in nonCustomerFacingCategories):
        dqMessage = f"You marked {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} as Customer Facing, but it uses the category {entryData['Category']} which is not allowed to be customer facing."
        summary.append(dqMessage)
        print(dqMessage)
        dqCheckValue = False
    if(entryData['Category'] in mustCustomerFacingCategories and not (entryData['Customer Facing Virtual'] == 'Yes' or entryData['Customer Facing Onsite'] == 'Yes')):
        dqMessage = f"You marked {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} as not Customer Facing, but it uses the category {entryData['Category']} which must be customer facing."
        summary.append(dqMessage)
        print(dqMessage)
        dqCheckValue = False
    if(type == 'ACCT'):
        if(categoryType not in ['ACCT', 'OPP/ACCT']):
            dqMessage = f"Category for Account OATS {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} was not valid. Make sure it is of type ACCT or OPP/ACCT."
            summary.append(dqMessage)
            print(dqMessage)
            dqCheckValue = False
        if(entryID not in listOfDummyAccounts):
            print('Applying Non Internal Account rules')
            if(categoryTypeDummy > -1):
                dqMessage = f"The Dummy Category is only valid for internal accounts. You tried to enter it for {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']}"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
            if(entryData['Category'] == 'Demonstration Development - OPP/ACCT'):
                dqMessage = f"Demonstration Development as a category is only allowed for an Internal Dummy Account or Opportunities. You tried to enter it for {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']}"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
        else:
            print('Applying Internal Account rules')
            internalAcademicSupportCats = [' DUMMY - Presentation / Meeting - ACCT', 'Preparation / Development - OPP/ACCT']
            internalMarketingSupportCats = [' DUMMY - Presentation / Meeting - ACCT', 'Preparation / Development - OPP/ACCT', 'Demonstration Development - OPP/ACCT', 'Travel to Customer Site - OPP/ACCT']
            internalPartnerSupportCats = ['Preparation / Development - OPP/ACCT', 'Partner Meeting, Planning / Review - OPP/ACCT']
            internalPropositionDevlopmentCats = [' DUMMY - Presentation / Meeting - ACCT', 'Preparation / Development - OPP/ACCT', 'Demonstration Development - OPP/ACCT']
            internalSalesCaEnablementCats = [' DUMMY - Presentation / Meeting - ACCT', 'Preparation / Development - OPP/ACCT', 'Travel to Customer Site - OPP/ACCT']
            internalBusinessPlanningCats = [' DUMMY - Presentation / Meeting - ACCT', 'Preparation / Development']
            internalTrainingCats = [' DUMMY - Receiving Training - ACCT']
            internalProductContentDevelopmentEnrichmentCats = ['Preparation / Development - OPP/ACCT', ' DUMMY - Presentation / Meeting - ACCT']
            if(entryID == '12356240' and entryData['Category'] not in internalAcademicSupportCats):
                dqMessage = f"For the {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} only the categories {internalAcademicSupportCats} are allowed"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
            elif(entryID == '12356250' and entryData['Category'] not in internalMarketingSupportCats):
                dqMessage = f"For the {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} only the categories {internalMarketingSupportCats} are allowed"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
            elif(entryID == '12356260' and entryData['Category'] not in internalPartnerSupportCats):
                dqMessage = f"For the {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} only the categories {internalPartnerSupportCats} are allowed"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
            elif(entryID == '12356270' and entryData['Category'] not in internalPropositionDevlopmentCats):
                dqMessage = f"For the {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} only the categories {internalPropositionDevlopmentCats} are allowed"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
            elif(entryID == '12356280' and entryData['Category'] not in internalSalesCaEnablementCats):
                dqMessage = f"For the {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} only the categories {internalSalesCaEnablementCats} are allowed"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
            elif(entryID == '13049830' and entryData['Category'] not in internalBusinessPlanningCats):
                dqMessage = f"For the {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} only the categories {internalBusinessPlanningCats} are allowed"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
            elif(entryID == '13049840' and entryData['Category'] not in internalTrainingCats):
                dqMessage = f"For the {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} only the categories {internalTrainingCats} are allowed"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
            elif(entryID == '13049850' and entryData['Category'] not in internalProductContentDevelopmentEnrichmentCats):
                dqMessage = f"For the {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} only the categories {internalProductContentDevelopmentEnrichmentCats} are allowed"
                summary.append(dqMessage)
                print(dqMessage)
                dqCheckValue = False
    elif(type == 'OPP'):
        if(categoryType not in ['OPP', 'OPP/ACCT']):
            dqMessage = f"Category for Opportunity OATS {entryData['ID (Please curate in the List sheet)']} on the {entryData['Date (e.g. 14Mar22)']} was not valid. Make sure it is of type OPP or OPP/ACCT."
            summary.append(dqMessage)
            print(dqMessage)
            dqCheckValue = False
    return dqCheckValue

# Iterate over the Accounts Level OATS
for index, row in dfAccounts.iterrows():
    # Extract the ID from the String
    idStart = row['ID (Please curate in the List sheet)'].find('||') + 4
    id = row['ID (Please curate in the List sheet)'][idStart: -1]
    dqCheck = dataQualityCheck('ACCT', row, id)
    if(dqCheck):
        print(f"Adding a {row['Duration']} OATS entry on the Account: {row['ID (Please curate in the List sheet)']}")
        print(f"Adding the following Category {row['Category']}")
        print(f"Adding the following Value Proposition {row['Value Proposition']}")
        driver.get(f"{baselineURL}accounts/{id}/oats")
        time.sleep(10)
        # Find the Add Button on the OATS page
        addElement = driver.find_element(By.XPATH, '//*[@id="orion-site-content"]/orion-account-detail/div/orion-account-oats/orion-section/section/mat-card/div[1]/mat-card-header/orion-section-header/button[1]')
        addElement.click()
        time.sleep(5)
        # From here the interaction is key based, as the elements shift their IDs
        actions = ActionChains(driver)
        # Tab to the Date Picker
        actions = actions.send_keys(Keys.TAB)
        actions = actions.send_keys(row['Date (e.g. 14Mar22)'])
        # Tab to the Category Field
        actions = actions.send_keys(Keys.TAB)
        actions = actions.send_keys(Keys.TAB)
        actions = actions.pause(1)
        actions = actions.send_keys(Keys.ENTER)
        actions = actions.pause(1)
        actions = actions.send_keys(row['Category'])
        actions = actions.pause(1)
        actions = actions.send_keys(Keys.ARROW_DOWN)
        actions = actions.send_keys(Keys.ENTER)
        # One Tab to Duration Field
        actions = actions.send_keys(Keys.TAB)
        for i in range(durationDictionary[row['Duration']]):
            actions = actions.send_keys(Keys.ARROW_DOWN)
        # One Tab to Value Prop Field
        actions = actions.send_keys(Keys.TAB)
        actions = actions.pause(1)
        actions = actions.send_keys(Keys.ENTER)
        actions = actions.pause(1)
        actions = actions.send_keys(row['Value Proposition'])
        actions = actions.pause(1)
        actions = actions.send_keys(Keys.ARROW_DOWN)
        actions = actions.send_keys(Keys.ENTER)
        # Tab for the Checkboxes
        actions = actions.send_keys(Keys.TAB)
        if(row['Customer Facing Virtual'] == 'Yes'):
            actions = actions.send_keys(Keys.SPACE)
        elif(row['Customer Facing Onsite'] == 'Yes'):
            actions = actions.send_keys(Keys.TAB)
            actions = actions.send_keys(Keys.SPACE)
        else:
            actions = actions.send_keys(Keys.TAB)
        actions = actions.send_keys(Keys.TAB)
        if(row['Demonstration Given'] == 'Yes'):
            actions = actions.send_keys(Keys.SPACE)
        actions = actions.send_keys(Keys.TAB)
        if(row['Partner Present'] == 'Yes'):
            actions = actions.send_keys(Keys.SPACE)
        actions = actions.send_keys(Keys.TAB)
        if(row['Interregional Collab'] == 'Yes'):
            actions = actions.send_keys(Keys.SPACE)
        # Tab to Save Button
        actions = actions.send_keys(Keys.TAB)
        actions = actions.send_keys(Keys.TAB)
        actions = actions.send_keys(Keys.ENTER)
        actions.perform()
        # Check for success
        try:
            element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".toast-success")))
            # Update the Done cell value to mark row as enterd
            dataSheet.cell(row = row['index'] + 2, column = 12).value = 'X'
            print('OATS successfully added')
            print(f"You can review the entry here: {baselineURL}accounts/{id}/oats")
            saveExcel()
        except:
            print('Unable to enter this OATS - please check the entry')
    else:
        print('DQ Check failed, the entry is skipped. Please fix the reported issue')

# Inform user about how many Opportunity level OATS will be entered
if(len(dfOpps.index) == 0):
    print('No OATS for oppurtunities to enter.')
else:
    print(f"Adding {len(dfOpps)} opportunity OATS.")

# Iterate over the Opportunity Level OATS
for index, row in dfOpps.iterrows():
    # Extract the ID from the String
    idStart = row['ID (Please curate in the List sheet)'].find('||') + 4
    id = row['ID (Please curate in the List sheet)'][idStart: -1]
    dqCheck = dataQualityCheck('OPP', row, id)
    if(dqCheck):
        print(f"Adding a {row['Duration']} OATS entry on the Opp: {row['ID (Please curate in the List sheet)']}")
        print(f"Adding the following Category {row['Category']}")
        # Example https://orion.sas.com/opportunities/15764350/oats
        driver.get(f"{baselineURL}opportunities/{id}/oats")
        time.sleep(10)
        # Find the Add Button on the OATS page
        addElement = driver.find_element(By.XPATH, '//*[@id="orion-site-content"]/orion-opportunity-detail/div/orion-opportunity-oats/orion-section/section/mat-card/div[1]/mat-card-header/orion-section-header/button[1]')
        addElement.click()
        time.sleep(5)
        # Find the Date Input on the OATS page
        dateElement = driver.find_element(By.XPATH, '//input[@data-mat-calendar="mat-datepicker-0"]')
        driver.execute_script('arguments[0].value = "";', dateElement)
        dateElement.send_keys(row['Date (e.g. 14Mar22)'])
        # From here the interaction is key based, as the elements shift their IDs
        actions = ActionChains(driver)
        # Two Tabs to get from Date Field to Category Field
        actions = actions.send_keys(Keys.TAB)
        actions = actions.send_keys(Keys.TAB)
        actions = actions.pause(1)
        actions = actions.send_keys(Keys.ENTER)
        actions = actions.pause(1)
        actions = actions.send_keys(row['Category'])
        actions = actions.pause(1)
        actions = actions.send_keys(Keys.ARROW_DOWN)
        actions = actions.send_keys(Keys.ENTER)
        # One Tab to Duration Field
        actions = actions.send_keys(Keys.TAB)
        for i in range(durationDictionary[row['Duration']]):
            actions = actions.send_keys(Keys.ARROW_DOWN)
        # Tab for the Checkboxes
        actions = actions.send_keys(Keys.TAB)
        if(row['Customer Facing Virtual'] == 'Yes'):
            actions = actions.send_keys(Keys.SPACE)
        elif(row['Customer Facing Onsite'] == 'Yes'):
            actions = actions.send_keys(Keys.TAB)
            actions = actions.send_keys(Keys.SPACE)
        else:
            actions = actions.send_keys(Keys.TAB)    
        actions = actions.send_keys(Keys.TAB)
        if(row['Demonstration Given'] == 'Yes'):
            actions = actions.send_keys(Keys.SPACE)
        actions = actions.send_keys(Keys.TAB)
        if(row['Partner Present'] == 'Yes'):
            actions = actions.send_keys(Keys.SPACE)
        actions = actions.send_keys(Keys.TAB)
        if(row['Interregional Collab'] == 'Yes'):
            actions = actions.send_keys(Keys.SPACE)
        # Tab to Save Button
        actions = actions.send_keys(Keys.TAB)
        actions = actions.send_keys(Keys.TAB)
        actions = actions.send_keys(Keys.ENTER)
        actions.perform()
        # Check for success
        try:
            element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".toast-success")))
            # Update the Done cell value to mark row as enterd
            dataSheet.cell(row = row['index'] + 2, column = 12).value = 'X'
            print('OATS successfully added')
            print(f"You can review the entry here: {baselineURL}opportunities/{id}/oats")
            saveExcel()
        except:
            print('Unable to enter this OATS - please check the entry')
    else:
        print('DQ Check failed, the entry is skipped. Please fix the reported issue')


# Close the Excel file
wb.close()
# Ending the browser session
driver.quit()

print('*************DQ-Issue-SUMMARY**************')
if(len(summary) == 0):
    print('No DQ issues were detected, please still check the DQ Report from time to time to make sure.')
for entry in summary:
    print(entry)
print('**********************************')